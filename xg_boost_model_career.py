# -*- coding: utf-8 -*-
"""
Created on Tue Jan 21 16:17:26 2020

@author: DBDA
"""

import pandas as pd

#df = pd.read_excel("newprojectdataset.xlsx")

df = pd.read_excel("C:\\Users\\dbda\\Desktop\\book1.xlsx")
dum_df = pd.get_dummies(df)
#dum_df = dum_df.drop('Class_Benign', axis=1)

from sklearn.model_selection import train_test_split 
from sklearn.metrics import confusion_matrix
from sklearn.metrics import classification_report, accuracy_score
#from sklearn.ensemble import BaggingClassifier

X = dum_df.iloc[:,0:20]
y = dum_df.iloc[:,20]

X_train, X_test, y_train, y_test = train_test_split(X, y,test_size = 0.3, 
                                                    random_state=2018,
                                                    stratify=y)
#
## Default: Tree Classifier
#model_rf = BaggingClassifier(random_state=1211,oob_score=True,
#                             max_features=X.shape[1],
#                             n_estimators=50)
#
##OR for any other model bagging
#from sklearn.linear_model import LogisticRegression
#
#model_rf = BaggingClassifier(base_estimator = LogisticRegression() ,
#                             random_state=1211,oob_score=True,
#                             max_features=X.shape[1])
#                             
#model_rf.fit( X_train , y_train )
#
#print("Out of Bag Score = " + "{:.4f}".format(model_rf.oob_score_))
#
#y_pred = model_rf.predict(X_test)
#
#print(confusion_matrix(y_test, y_pred))
#print(classification_report(y_test, y_pred))
##print(accuracy_score(y_test, y_pred))
## Gradient BOOST
#from sklearn.model_selection import train_test_split 
#from sklearn.metrics import confusion_matrix, classification_report
#from sklearn.metrics import accuracy_score
#from sklearn.ensemble import GradientBoostingClassifier
#
## Create training and test sets
#X_train, X_test, y_train, y_test = train_test_split(X, y, test_size = 0.3, 
#                                                    random_state=2018)
#
#clf = GradientBoostingClassifier(random_state=1200)
#clf.fit(X_train,y_train)
#
#y_pred = clf.predict(X_test)
#
#
##inputdata = pd.read_excel("C:\\Users\\dbda\\Desktop\\newprojectdataset.xlsx")
##inputdata=pd.get_dummies(inputdata)
##inputtest=inputdata.iloc[90000:,:20]
#
##predictedout=clf.predict(inputtest)
##prediction = pd.DataFrame(predictedout, columns=['predictions']).to_csv('C:\\Users\\dbda\\Desktop\\prediction120000.csv')
#
#print(confusion_matrix(y_test, y_pred))
#print(classification_report(y_test, y_pred))
#print(accuracy_score(y_test,y_pred))

#----------------------------------------
#XGBOOST 
from xgboost import XGBClassifier

# Create training and test sets
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size = 0.3, 
                  
                                                    random_state=2018)

dt=pd.DataFrame(np.array([5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5]))
#dt.valuestolist()
import openpyxl 

wb_obj = openpyxl.load_workbook("test.xlsx") 
sheet_obj = wb_obj.active
 
cell_obj = sheet_obj.cell(row = 1, column = 1) 
print(cell_obj.value)
max_row=sheet_obj.max_row
print(max_row)
max_col=sheet_obj.max_column
for i in range(1, max_col + 1): 
    cell_obj = sheet_obj.cell(row = max_row, column = i) 
    print(cell_obj.value, end = " ") 
for i in dt.values:
    print(i)
#f=open("test.xlsx","a",sep=" ")
#for i in f:
#    print(i)
    
#    
#dt=pd.DataFrame(np.array([5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5]))
#    
clf = XGBClassifier(random_state=2000,max_depth=3)
clf.fit(X_train,y_train)

import numpy as np
#pd.DataFrame
#dt=pd.DataFrame(np.array([2,3,2,4,1,5,4,2,3,4,3,3,2,1,4,4,4,3,4,4]))
#
#dt=pd.DataFrame(np.array([5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5]))
#dt=dt.T
#pd.DataFrame('q1':3,'q2':4,'q3':5,'q4':5,'q5':3,'q6':2,'q7':1,'q8':5,'q9':2,'q10':3,'q11':3,'q12':2,'q13':4,'q14':3,'q15':5,'q16':3,'q17':3,'q18':1,'q19':3,'q20':2

##with open 
#prediction = pd.DataFrame(dt, columns=['q1','q2','q3','q4','q5','q6','q7','q8','q9','q10','q11','q12','q13','q14','q15','q16','q17','q18','q19','q20']).to_csv('C:\\Users\\dbda\\Desktop\\topredict.csv')
#
##col=['q1','q2','q3','q4','q5','q6','q7','q8','q9','q10','q11','q12','q13','q14','q15','q16','q17','q18','q19','q20']
#new=pd.DataFrame(data=[5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5],columns=col,axis=1)
dtest=pd.read_excel(r"F:\\project\\webapp_making\\Test1.xlsx")

#y_pred = clf.predict(X_test)
y_pred = clf.predict(dtest)
print(y_pred)

y_pred = clf.predict(X_test)
print(confusion_matrix(y_test, y_pred))
print(classification_report(y_test, y_pred))
print(accuracy_score(y_test,y_pred))